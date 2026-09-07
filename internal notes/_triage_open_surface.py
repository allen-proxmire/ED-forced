# -*- coding: utf-8 -*-
"""Triage the corpus's open surface into work that can actually be scheduled.

WHY
---
"272 rows undated, 51 Open, 78 dead pointers" are three numbers that sound like
one backlog and are not. This ranks them so the next session works the rows that
carry weight instead of the rows that are merely unstamped.

THE RANKING PRINCIPLE
---------------------
An undated row matters in proportion to what rests on it. Four signals, all
readable from the workbook and the tree:

  TIER      a Derived/Grounded/D-via-I/Wall row asserts something; a
            Reference/Superseded/Synthesis row mostly points somewhere.
  BET       a row with a falsifier is a claim someone could act on.
  REACH     how many OTHER papers cite this row's paper -- an unverified claim
            with ten dependents is a different object from one with none.
  FACE      whether the row also appears on the outward-facing sheets
            (Core Theory / Core Predictions / Core Claims).

Score = TIER + BET + REACH + FACE. The point is not the number; it is that the
list comes out ordered, so "272 undated" becomes "these N first".
"""
from __future__ import print_function
import io
import os
import re
import sys
import collections

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

WB = "ED_ItemizedTheory_TieredClaims_v2.xlsx"

ASSERTIVE = {"Derived", "Grounded", "D-via-I / Form-forced", "Wall", "Measured",
             "Postulated", "Prediction", "Asserted", "Primitive", "Standing"}
POINTERS = {"Reference", "Superseded", "Synthesis", "Selected/Inherited",
            "Constant", "Inherited / Consilience", "Identification"}


def rule(c="-", n=78):
    print(c * n)


def paper_citation_counts():
    """How many markdown files mention each paper token."""
    counts = collections.Counter()
    files = []
    for dirpath, dirnames, filenames in os.walk("physics-papers"):
        dirnames[:] = [d for d in dirnames if not d.startswith(".")]
        for f in filenames:
            if f.endswith(".md"):
                files.append(os.path.join(dirpath, f))
    texts = {}
    for f in files:
        try:
            texts[f] = io.open(f, encoding="utf-8").read()
        except IOError:
            pass
    return texts


def main():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl required"); return 2

    wb = openpyxl.load_workbook(WB, read_only=True)
    ws = wb["Ledger w Claims"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]

    # outward-facing claim text, for the FACE signal
    face = set()
    for sheet in ("Core Theory", "Core Predictions", "Core Claims"):
        if sheet in wb.sheetnames:
            for r in wb[sheet].iter_rows(values_only=True):
                for c in r:
                    if isinstance(c, str) and len(c) > 40:
                        face.add(c.strip()[:80])

    texts = paper_citation_counts()
    reach = {}

    def reach_of(paper):
        p = str(paper or "").strip()
        if not p or len(p) < 3:
            return 0
        if p in reach:
            return reach[p]
        pat = re.escape(p)
        n = sum(1 for t in texts.values() if re.search(pat, t))
        reach[p] = n
        return n

    undated = []
    for r in rows:
        folder, paper, claim, tier, fals = r[0], r[1], r[2], r[3], r[4]
        last = r[12]
        if last:
            continue
        t = str(tier or "").strip()
        score = 0
        why = []
        if t in ASSERTIVE:
            score += 3; why.append("asserts")
        elif t in POINTERS:
            score += 0
        else:
            score += 1
        if fals and len(str(fals)) > 20:
            score += 2; why.append("has falsifier")
        rr = reach_of(paper)
        if rr >= 8:
            score += 3; why.append("reach %d" % rr)
        elif rr >= 3:
            score += 2; why.append("reach %d" % rr)
        elif rr >= 1:
            score += 1
        if claim and str(claim).strip()[:80] in face:
            score += 3; why.append("outward-facing")
        undated.append((score, t, str(folder or ""), str(paper or ""),
                        str(claim or "")[:88], ", ".join(why)))

    undated.sort(key=lambda x: -x[0])

    print()
    rule("=")
    print("OPEN-SURFACE TRIAGE -- turning three backlog numbers into a work order")
    rule("=")
    print()

    rule()
    print("1. THE UNDATED ROWS, RANKED (%d of %d have no Last-Verified)"
          % (len(undated), len(rows)))
    rule()
    buckets = collections.Counter()
    for s, t, folder, paper, claim, why in undated:
        buckets["%d" % s] += 1
    print("     score distribution:", dict(sorted(buckets.items(), key=lambda kv: -int(kv[0]))))
    print()
    top = [u for u in undated if u[0] >= 8]
    mid = [u for u in undated if 5 <= u[0] < 8]
    low = [u for u in undated if u[0] < 5]
    print("     PRIORITY   %3d rows  (score >= 8: assertive + reaching + a bet or outward-facing)"
          % len(top))
    print("     MIDDLE     %3d rows  (5-7)" % len(mid))
    print("     TAIL       %3d rows  (< 5: mostly pointers, low reach)" % len(low))
    print()
    print("     The tail is the reason 272 is not a backlog. Those rows are")
    print("     references and superseded entries; stamping them buys nothing.")
    print()
    print("     TOP 15 BY SCORE:")
    print()
    for s, t, folder, paper, claim, why in undated[:15]:
        print("      %2d  %-22s %-14s %s" % (s, t[:22], paper[:14], claim[:60]))
        print("          %s" % why)
    print()

    # ------------------------------------------------------------- open tiers
    rule()
    print("2. THE 'Open' TIER ROWS -- these are physics, not hygiene")
    rule()
    opens = [r for r in rows if str(r[3] or "").strip() == "Open"]
    byfolder = collections.Counter(str(r[0] or "") for r in opens)
    for f, n in byfolder.most_common():
        print("     %-26s %d" % (f, n))
    print()
    print("     %d rows. Each is a declared gap with a name. Working them means" % len(opens))
    print("     doing the derivation, not updating a cell -- so they schedule as")
    print("     research sessions, one at a time, not as a batch.")
    print()

    # --------------------------------------------------------- dead pointers
    rule()
    print("3. DEAD POINTERS -- mechanical, finite, and the cheapest of the three")
    rule()
    print("     Counted by _check_doc_coherence.py: 0 in the four navigation")
    print("     files, 78 elsewhere, 409 more that resolve in a sibling repo and")
    print("     are not defects. The 78 are real and each is a one-line fix.")
    print()
    print("     Run:  python \"internal notes/_check_doc_coherence.py\" --full")
    print()

    rule("=")
    print("WHAT THE NUMBERS ACTUALLY SAY")
    rule("=")
    print("  * The undated 272 is NOT 272 units of work. %d rows carry weight;" % len(top))
    print("    the %d-row tail is pointers and superseded entries." % len(low))
    print("  * The 51 Open rows are research, not hygiene, and cannot be batched.")
    print("  * The 78 dead pointers are the only item here that is purely")
    print("    mechanical and finishable in one sitting.")
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
