# -*- coding: utf-8 -*-
"""Reconcile every item from both review rounds against the files.

AP asked whether the reviews are resolved. Earlier today I answered that question
from memory and was wrong. So this enumerates each item and tests it.
"""
import io, os, sys
sys.stdout.reconfigure(encoding="utf-8")
os.chdir(r"C:\Users\allen\GitHub\ED Generative")
import openpyxl

REPORT = io.open("ED_UnifiedFramework_Report.md", encoding="utf-8").read()
PAPER = io.open("physics-papers/gravity/Paper_a0z_MONDScaleTracksHubbleRate.md",
                encoding="utf-8").read()
ML = io.open("physics-papers/predictions/ED_Master_Predictions_List.md", encoding="utf-8").read()
wb = openpyxl.load_workbook("ED_ItemizedTheory_TieredClaims_v2.xlsx")
ws = wb["Ledger w Claims"]


def cell(r, c):
    return str(ws.cell(r, c).value or "")


# (round, reviewer, item, test) -- test returns True if resolved
ITEMS = [
    # ---------------- ROUND A -------------------------------------------------
    ("A", "claude", "alpha = 1.18 mislabelled as the survey's measurement",
     lambda: "converted by us from their linear fit" in PAPER and "8b" in PAPER),
    ("A", "claude", "LCDM (Magneticum) also predicts apparent a0 evolution",
     lambda: "Magneticum" in PAPER and "Magneticum" in REPORT),
    ("A", "claude", "the 'three baryonic systematics' paper -- verify or flag",
     lambda: "academia.edu/170831186" in PAPER),

    ("A", "gpt", "1. remove stale 'constructive sign is open'",
     lambda: "One residual is open and named: the constructive sign" not in REPORT),
    ("A", "gpt", "2. workbook Sign-Critical: structural vs measured",
     lambda: all("STRUCTURALLY SETTLED" in cell(r, 12) for r in (179, 191, 199))),
    ("A", "gpt", "3. 'closed sector' -> 'closed in the tested GR regime'",
     lambda: "It is a closed sector, but" not in REPORT
             and "closed in the GR regime that has been tested" in REPORT),
    ("A", "gpt", "4. qualify classical tests by the UFF prerequisite",
     lambda: "conditional on the still-unproven universal-free-fall normalisation" in REPORT),
    ("A", "gpt", "5. MOND residual list = threshold / high-acc / UFF, not 'sign'",
     # superseded by the physics audit, which found that list wrong in 2 of 3;
     # the item is resolved by the CORRECTED list, not the original one.
     lambda: "The MOND residuals — CORRECTED 2026-09-07 after a physics audit" in REPORT),
    ("A", "gpt", "6. shared Coh/Grad vocabulary, workbook == report",
     lambda: "measured corroboration CANDIDATE" in cell(191, 12)
             and "MEASURED CORROBORATION" in cell(191, 8)),

    ("A", "gemini", "no action items (arithmetic verification + endorsement)",
     lambda: True),

    # ---------------- ROUND B -------------------------------------------------
    ("B", "claude", "locate + read the systematics paper, report honestly",
     lambda: "Rusishvili" in PAPER and "5.3a" in PAPER),
    ("B", "claude", "check the three claims; say if they undercut the 30 sigma",
     lambda: "a0z_baryonic_systematics_check.py" in PAPER and "UNTESTED" in PAPER),
    ("B", "claude", "re-reduce the public catalogue yourself -- DECLINED by AP 2026-09-07,\n"
                    "                     recorded in the paper and ED_Research_Targets as a\n"
                    "                     decision rather than an oversight",
     lambda: "does not claim to have run that test, and does not intend to" in PAPER),

    ("B", "gpt", "1. fix the a0 paper ABSTRACT (was 4.4 sigma / ~1-2 sigma)",
     lambda: "4.4\u03c3" not in PAPER.split("## 1.")[0]
             and "2.3\u03c3" in PAPER.split("## 1.")[0]),
    ("B", "gpt", "2. workbook row 380 (P14 / regime / sign all stale)",
     lambda: "REWRITTEN 2026-09-07" in cell(380, 7)),
    ("B", "gpt", "3. workbook row 362 (P14 -> P-Quadratic-Strain)",
     # NB the cell legitimately QUOTES the old dependency while explaining the
     # re-scope ("was 'GIVEN P14'"), so test the live dependency, not the string.
     lambda: cell(362, 3).startswith("MOND = the off-diagonal")
             and "forced GIVEN P-Quadratic-Strain" in cell(362, 3)),
    ("B", "gpt", "4. KM-I status must not imply implementation settled",
     lambda: "canonicalization remains OPEN" in cell(120, 8)),
    ("B", "gpt", "5. currency date -> 2026-09-07",
     lambda: "currency pass 2026-09-07" in REPORT),
    ("B", "gpt", "6. (item 10) 'form-complete' is a COVERAGE claim",
     lambda: "COVERAGE claim" in REPORT),

    ("B", "gemini", "no action items (ledger alignment verification)",
     lambda: True),

    # ---------------- ROUND C (2026-09-07, after the hygiene + detector work) --
    ("C", "gpt", "1. a0 tier: Report must match the workbook's Derived -> Grounded",
     lambda: "form-forced conditional on the live-horizon reading" in REPORT
             and "the **evolution** is the claim here and it is derived" not in REPORT),
    ("C", "gpt", "2. Rusishvili status discrepancy -- FALSE ALARM, stale copy reviewed\n"
                 "                     (the current paper does not contain the sentence quoted)",
     lambda: "have not been able to independently confirm" not in PAPER
             and "academia.edu/170831186" in PAPER),
    ("C", "gpt", "3. 'MOND killed' too broad",
     lambda: "killing the MOND picture" not in REPORT
             and "exclude constant-`a\u2080` MOND" in REPORT),
    ("C", "gpt", "4. workbook row 20 leads with STALE",
     lambda: not str(ws.cell(20, 8).value or "").startswith("STALE AS OF")),
    ("C", "claude", "5. §16 Bottom Line still quoted the superseded 4-sigma / 1-2 sigma",
     lambda: "2.3\u03c3 on a computed conversion budget" in REPORT),
    ("C", "claude", "6. Standard-Model comparison restated in four sections",
     lambda: REPORT.count("no worse than the Standard Model") <= 1
             and "is made once, in \u00a712" in REPORT),
    ("C", "gemini", "no action items (structural summary)", lambda: True),

    # ---------------- self-found, same round -----------------------------------
    ("C", "self", "physics audit: Report's MOND residual list was stale in 2 of 3",
     lambda: "CORRECTED 2026-09-07 after a physics audit" in REPORT),
    ("C", "self", "UFF narrowed to one exponent; 1/b_C repair refuted",
     lambda: "b_C^{\u22121/2}" in REPORT or "the same violation inverted" in REPORT),
]

print()
print("=" * 78)
print("REVIEW RECONCILIATION -- every item from both rounds, tested against the files")
print("=" * 78)
print()
open_items = []
for rnd, who, item, test in ITEMS:
    try:
        ok = bool(test())
    except Exception as e:
        ok = False
        item += "  [test error: %s]" % e
    mark = "RESOLVED" if ok else "OPEN"
    if not ok:
        open_items.append((rnd, who, item))
    print("  [%s] %-7s %-8s %s" % (rnd, who, mark, item))

print()
print("-" * 78)
print("  %d items, %d resolved, %d open" % (len(ITEMS), len(ITEMS) - len(open_items),
                                            len(open_items)))
print("-" * 78)
if open_items:
    print()
    for rnd, who, item in open_items:
        print("  OPEN  [%s/%s] %s" % (rnd, who, item))
print()
sys.exit(0)
