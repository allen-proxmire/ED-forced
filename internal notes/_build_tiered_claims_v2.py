# -*- coding: utf-8 -*-
"""Build v2 of the tiered-claims workbook: add the `Rests On` column.

Ledger #122 found the defect in the GR/QFT block is a LINKAGE failure, not
concealment: the postulates are declared in their own rows, correctly tiered
`Postulated / live`, and the `Derived` rows that depend on them all carry
`Postulate: None`.  The fix is to make the dependency explicit.

WHAT THIS WRITES.  A NEW file, `..._v2.xlsx`.  The original is never touched.
Every sheet is copied through; `Ledger w Claims` gains three columns:

  Rests On       what the claim depends on beyond P01-P13.  Filled for the 55
                 `Derived` rows -- the tier under scrutiny -- at three
                 confidence levels, each labelled:
                   VERIFIED     read against the paper on 2026-09-06 (#122)
                   CANDIDATE    the paper declares this postulate; whether THIS
                                claim needs it is unread
                   none found   the paper declares no censused postulate
  Sign-Critical  GPT's suggestion.  A claim whose conclusion flips with one
                 sign.  Seeded by keyword; every hit is a CANDIDATE for a human
                 to confirm and to attach an invariant to.
  Last Verified  the row's latest date, lifted out of the prose into a column
                 so staleness is sortable rather than regex-able.

AND SIX RE-TIERS, the ones verified in #122 against the papers' own words.  The
original tier is preserved in the `Rests On` cell so nothing is lost silently.

WHAT THIS DOES NOT DO.  It does not adjudicate the 30 CANDIDATE rows -- that
needs a human read of each paper, and pretending otherwise would reproduce the
exact error the audit found.  It does not re-date the 207 stale rows (C2) or
add the 87 missing postulates (C1); those change currency, not claims.

Run: python "internal notes/_build_tiered_claims_v2.py"
"""
import io
import os
import re
import shutil
import subprocess
import sys

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

SRC = "ED_ItemizedTheory_TieredClaims.xlsx"
DST = "ED_ItemizedTheory_TieredClaims_v2.xlsx"
PNAME = re.compile(r"P-[A-Za-z0-9][A-Za-z0-9\-]{2,}")
DATE = re.compile(r"2026-\d\d-\d\d")
FALSE_POSITIVES = {"P-CONSTRUCTION", "P-constructions", "P-definitional",
                   "P-imprint", "P-postulates"}
SIGNWORDS = re.compile(r"attractive|repulsive|constructive|destructive|"
                       r"inward|outward|parity|chirality|handedness|"
                       r"sign of|sign-|wrong sign|sign flip", re.I)
# Narrowed 2026-09-06: bare "positive"/"negative" fired on "non-negative
# additive scalar" and "structural-positive", which are not sign claims.

# --- Paper resolution -------------------------------------------------------
# The sheet's Paper column is free text: bare numbers ("029"), short names
# ("GR-I"), section refs ("Constants-paper §4"), pair refs ("089/093") and
# folder-local labels ("Cos-06").  A plain substring matcher on that put FIVE of
# 31 flags on the WRONG paper -- "Constants-paper §4" landed on Paper_004,
# "Cos-06" on Paper_006, "012.7" on Paper_001 -- because a stray digit matched.
# That is the same failure that made C4 report a false all-clear before, so it
# gets two independent defences: an explicit alias map, and a FOLDER GUARD.
# The guard alone catches all five: the sheet already records the folder, and a
# file resolved outside it is wrong by construction.

ALIAS = {
    "Constants-paper §4": ["constants-inherited/Constants_Ledger_and_G_Dimension.md"],
    "Cos-06":          ["cosmology/Paper_ED_Cos_06_InflationarySpectrum.md"],
    "DarkSector":      ["dark-sector/Paper_ED_DarkSector.md"],
    "MetricFromGraph": ["gravity/Paper_MetricFromTheGraph_ForcedTo3D.md"],
    "012.7":           ["qm-kinematics/Paper_012_7_AdjacencyBandwidth_Galilean.md"],
    "032.5":           ["gravity/Paper_032_5_FreeChainGeodesics.md"],
    "A1 (CommonCauseNotChannel)": ["substrate-evaluation/Paper_CommonCauseNotChannel_A1.md"],
    "B4 (ChargeAsTopology)":      ["substrate-evaluation/Paper_ChargeAsTopology_B4.md"],
    # pair refs: read BOTH papers, not whichever one happened to match first
    "089/093": ["foundations/Paper_089_V1Kernel.md", "foundations/Paper_093_KernelArrow_of_Time.md"],
    "090/094": ["foundations/Paper_090_V5Kernel.md", "foundations/Paper_094_ForwardCausality.md"],
    "034/036": ["gravity/Paper_034_DeepMOND.md", "gravity/Paper_036_MOND_FieldEquation.md"],
    "054/056": ["q-compute/Paper_054_UR1.md", "q-compute/Paper_056_ClassA_Wall.md"],
}

# The six verified in ledger #122, with the paper's own words as the warrant.
VERIFIED = {
    69: ("Derived", "Postulated",
         "P-Commitment-Linear. VERIFIED 2026-09-06 (#122): the claim text itself names "
         "the postulate. Self-contradicting as filed."),
    63: ("Derived", "Grounded",
         "P-Commitment-Linear. VERIFIED 2026-09-06 (#122): GR-I preamble 6 -- 'the lapse "
         "derivation rests on a stated postulate (P-Commitment-Linear, section 2) ... the "
         "load-bearing structural commitment selecting the Einstein branch'."),
    64: ("Derived", "Grounded",
         "P-Commitment-Linear. VERIFIED 2026-09-06 (#122): same lapse chain as row 63."),
    65: ("Derived", "D-via-I / Form-forced",
         "Lovelock uniqueness (conditional) + inherited kappa, Lambda. VERIFIED 2026-09-06 "
         "(#122): GR-II preamble 2 makes the form conditional; preamble 7 -- 'kappa = 8 pi G "
         "and Lambda are inherited, not derived'. Form derived, values inherited."),
    80: ("Derived", "Grounded",
         "P-YM-Action-Coarse-Graining. VERIFIED 2026-09-06 (#122): Paper_019 section 3.4 -- "
         "'standard variational derivation from the P-postulated action ... the action itself "
         "is constructed via P-YM-Action-Coarse-Graining'."),
    81: ("Derived", "Grounded",
         "P-Gap-Coercivity, P-Profile-Rescaling, P-Quartic-Sign. VERIFIED 2026-09-06 (#122): "
         "Paper_023 -- 'Mass-gap mechanism with P-Gap-Coercivity, P-Profile-Rescaling, "
         "P-Quartic-Sign (Paper_021)'."),
}


def census_names():
    out = subprocess.run([sys.executable, "internal notes/_census_postulates.py", "--list"],
                         capture_output=True, text=True, encoding="utf-8").stdout
    return set(re.findall(r"^  (P-[A-Za-z0-9][A-Za-z0-9\-]*)", out, re.M))


def paper_files():
    out = {}
    for dp, _dn, fn in os.walk("physics-papers"):
        for f in fn:
            if not f.endswith(".md"):
                continue
            path, stem = os.path.join(dp, f), f[:-3]
            out.setdefault(stem, path)
            m = re.match(r"Paper_([0-9]+(?:_[0-9]+)?)", stem)
            if m:
                out.setdefault(m.group(1), path)
                out.setdefault(m.group(1).lstrip("0"), path)
            m2 = re.match(r"Paper_((?:GR|KM|MS|RQM|ED)[-_][A-Za-z0-9\-]+)", stem)
            if m2:
                out.setdefault(m2.group(1).replace("_", "-"), path)
    return out


def resolve(name, folder, files):
    """Return (list_of_paths, how) where how is EXACT / ALIAS / FUZZY / UNRESOLVED.

    A fuzzy hit whose directory disagrees with the sheet's own Folder column is
    REJECTED rather than returned: it is a digit collision, and guessing there
    is exactly what produced the five wrong flags."""
    if name in ALIAS:
        return [os.path.join("physics-papers", p.replace("/", os.sep))
                for p in ALIAS[name]], "ALIAS"
    if name in files:
        return [files[name]], "EXACT"
    want = (folder or "").strip().lower()
    if name and len(name) > 2:
        for k, v in sorted(files.items()):
            if name in k or k in name:
                got = os.path.basename(os.path.dirname(v)).lower()
                if want and got != want:
                    continue                    # folder guard: wrong neighbourhood
                return [v], "FUZZY"
    return [], "UNRESOLVED"


def main():
    import openpyxl
    if not os.path.exists(SRC):
        print("source workbook not found: %s" % SRC)
        return 1
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["Ledger w Claims"]

    hdr = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(hdr) if h}
    base = ws.max_column
    cRest, cSign, cVer = base + 1, base + 2, base + 3
    ws.cell(row=1, column=cRest, value="Rests On")
    ws.cell(row=1, column=cSign, value="Sign-Critical")
    ws.cell(row=1, column=cVer, value="Last Verified")

    cen = census_names()
    files = paper_files()
    n_ver = n_cand = n_none = n_sign = n_dated = n_unres = 0

    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, base + 1)]
        alltxt = " ".join(str(v) for v in vals if v)
        if not alltxt.strip():
            continue
        tier = str(ws.cell(row=r, column=idx["Tier"]).value or "")
        paper = str(ws.cell(row=r, column=idx["Paper"]).value or "")
        claim = str(ws.cell(row=r, column=idx["Claim"]).value or "")

        ds = DATE.findall(alltxt)
        if ds:
            ws.cell(row=r, column=cVer, value=max(ds))
            n_dated += 1

        if SIGNWORDS.search(claim):
            ws.cell(row=r, column=cSign,
                    value="CANDIDATE - confirm, and record the invariant that fixes the sign")
            n_sign += 1

        if r in VERIFIED:
            old, new, why = VERIFIED[r]
            ws.cell(row=r, column=idx["Tier"], value=new)
            ws.cell(row=r, column=cRest,
                    value="%s  [re-tiered %s -> %s, 2026-09-06, gravity ledger #122]" % (why, old, new))
            n_ver += 1
            continue

        if not tier.lower().startswith("derived"):
            continue

        folder = str(ws.cell(row=r, column=idx["Folder"]).value or "")
        paths, how = resolve(paper, folder, files)
        if how == "UNRESOLVED":
            ws.cell(row=r, column=cRest,
                    value="UNRESOLVED - the sheet's Paper name '%s' maps to no file under "
                          "physics-papers/%s. NOT checked. Fix the name, then re-run. "
                          "(2026-09-06)" % (paper, folder.lower()))
            n_unres += 1
            continue
        decl = set()
        for p in paths:
            if os.path.exists(p):
                decl |= set(PNAME.findall(io.open(p, encoding="utf-8", errors="replace").read()))
        decl = (decl - FALSE_POSITIVES) & cen
        read = ", ".join(os.path.basename(p) for p in paths)
        if decl:
            ws.cell(row=r, column=cRest,
                    value="CANDIDATE - %s declares %s. Whether THIS claim needs it is UNREAD; "
                          "a human must check. (C4 flag, 2026-09-06; resolved %s)"
                          % (read, ", ".join(sorted(decl)[:4]), how))
            n_cand += 1
        else:
            ws.cell(row=r, column=cRest,
                    value="none found - %s declares no censused postulate (mechanical check "
                          "only, 2026-09-06; resolved %s)" % (read, how))
            n_none += 1

    wb.save(DST)
    print("Wrote %s (original untouched)\n" % DST)
    print("  Derived rows re-tiered, VERIFIED against the paper : %d" % n_ver)
    print("  Derived rows marked CANDIDATE (needs a human read)  : %d" % n_cand)
    print("  Derived rows with no censused postulate found       : %d" % n_none)
    print("  Derived rows UNRESOLVED (name maps to no file)      : %d" % n_unres)
    print("  Sign-Critical candidates seeded                     : %d" % n_sign)
    print("  Rows given a Last Verified date                     : %d" % n_dated)
    print("""
  The CANDIDATE rows are the work that remains, and they are deliberately not
  adjudicated here: deciding them without reading each paper would reproduce
  exactly the error the audit found.
""")
    return 0


if __name__ == "__main__":
    sys.exit(main())
