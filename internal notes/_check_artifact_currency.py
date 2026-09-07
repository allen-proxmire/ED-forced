# -*- coding: utf-8 -*-
"""Verify the three outward-facing artifacts are current and agree with each other.

WHY THIS EXISTS
---------------
On 2026-09-07 I twice reported documents as current when they were not: once
because a canonical list was stale and fed an error into a new paper (ledger
#151), and once because I read three external reviews, acted on one, and then
described "the external review" as addressed (#153). Both times the thing that
caught it was AP asking, not a check.

Neither failure is exotic. Both are the same shape: a claim about a document's
state, made from memory of what I did rather than from the document. This script
reads the documents.

THE THREE ARTIFACTS
    ED_UnifiedFramework_Report.md          the synthesis
    ED_ItemizedTheory_TieredClaims_v2.xlsx the ledger of record
    physics-papers/gravity/Paper_a0z_...md the standalone prediction paper

FOUR CLASSES OF CHECK
    A  STALE     text that a fix was supposed to remove and did not
    B  APPLIED   text a fix was supposed to add, verified present
    C  AGREE     shared numbers that must match across artifacts
    D  FRESH     PDFs newer than sources; workbook counts match its own rows

Run with --selftest to confirm the checks can actually fail: it plants a stale
string in memory and asserts the checker flags it.

Exit 0 clean, 1 on any finding.
"""
from __future__ import print_function
import io
import os
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

REPORT = "ED_UnifiedFramework_Report.md"
PAPER = "physics-papers/gravity/Paper_a0z_MONDScaleTracksHubbleRate.md"
WORKBOOK = "ED_ItemizedTheory_TieredClaims_v2.xlsx"
REVIEW = "physics-papers/gravity/Note_a0z_Paper_AdversarialReview_2026-09-06.md"

# --------------------------------------------------------------------------- A
# Each entry: (file, regex, what it means, which ledger item should have killed it)
STALE = [
    (REPORT, r"It is a closed sector, but",
     "'a closed sector' contradicts the edge-list beneath it", "GPT item 3 / #153"),
    (REPORT, r"One residual is open and named: the constructive sign",
     "the constructive sign is structurally settled, not open", "GPT item 1 / #153"),
    (REPORT, r"softening to ~1–2σ once one-survey systematics are folded in",
     "the '~1-2 sigma' was asserted; the computed figure is 2.3", "#152"),
    (REPORT, r"a direct fit gives",
     "alpha = 1.18 is an ED-side recast, not the survey's fit", "#151"),
    (PAPER, r"\*\*I — measurement\*\* \| same source",
     "audit row 8 mislabelled our conversion as the survey's measurement", "#151"),
    (PAPER, r"tension softens to ~1–2σ under a realistic error budget",
     "step 9's special pleading, replaced by the computed budget", "#152"),
    (PAPER, r"untouched here — this paper does not test it",
     "LCDM is not untouched; Magneticum predicts apparent evolution", "#151"),
    ("physics-papers/predictions/ED_Master_Predictions_List.md",
     r"the implied power is \*\*\$\\alpha = 1\.18",
     "canonical row must attribute alpha to ED, not the survey", "#151"),
    ("physics-papers/predictions/ED_Master_Predictions_List.md",
     r"α=1 is \*dimensionally forced\*",
     "alpha=1 is mechanism-forced; ED carries l_P", "#151"),
    ("physics-papers/predictions/22_Ways_to_Kill_Event_Density.md", r"a direct fit gives",
     "alpha provenance", "#151"),
    ("physics-papers/predictions/Paper_101_FalsificationRegister.md", r"a direct fit gives",
     "alpha provenance", "#151"),
    ("essays/Paper_SelectedFormalisms_I_Gravity.md", r"A direct fit gives",
     "alpha provenance", "#151"),

    # --- added 2026-09-07 after this checker MISSED the stale abstract ---------
    # The original list keyed on section 5.3's exact wording, so the abstract's
    # different phrasing of the SAME stale claim sailed through and an external
    # reviewer found it. A pattern list only catches the phrasings you thought of.
    # These match the NUMBERS instead, which are phrasing-independent: any
    # superseded figure anywhere in the paper is now a finding regardless of how
    # the sentence around it is worded.
    (PAPER, r"4\.4σ", "superseded tension figure (now 2.3σ) anywhere in the paper", "#155"),
    (PAPER, r"1–2σ", "superseded softening claim anywhere in the paper", "#155"),
    (PAPER, r"α = 1\.18 ± 0\.04 \(stat\)\*\*\. MOND",
     "abstract quoting the old conversion as the result", "#155"),
    (REPORT, r"currency pass 2026-09-06", "currency date behind the content", "#155"),

    # --- added 2026-09-07 after the checker missed the Report's Bottom Line ----
    # Second time this blind spot fired: a per-file pattern list inherits the
    # blind spot of whoever wrote it. The paper got number-matching after the
    # abstract miss; the Report did not, and a reviewer found the same defect in
    # §16. Match the NUMBERS here too.
    (REPORT, r"1\u20132\u03c3 under systematics",
     "Bottom Line quoting the superseded softening", "#161"),
    (REPORT, r"nominal ~4\u03c3", "superseded nominal tension figure", "#161"),
    (REPORT, r"killing the MOND picture",
     "claims more than excluding constant-a0 MOND", "#161"),
    (REPORT, r"the \*\*evolution\*\* is the claim here and it is derived",
     "tier out of sync with the workbook's Derived -> Grounded", "#161"),
]

# --------------------------------------------------------------------------- B
APPLIED = [
    (REPORT, "Magneticum", "the LCDM rival is named", "#151"),
    (REPORT, "closed in the GR regime that has been tested", "closed-sector fix", "GPT 3"),
    (REPORT, "The MOND residuals, stated as the three they actually are", "residual list", "GPT 5"),
    (REPORT, "conditional on the still-unproven universal-free-fall normalisation",
     "abstract conditioned on UFF", "GPT 4"),
    (REPORT, "2.3σ on a computed conversion budget", "computed tension", "#152"),
    (PAPER, "Magneticum", "LCDM rival in the four-way table", "#151"),
    (PAPER, "a0z_powerlaw_refit.py", "the refit is cited and reproducible", "#152"),
    (PAPER, "5.3a", "the unverified rebuttal is recorded as a check", "#151"),
    (PAPER, "5.3b", "the shape comparison exists", "#152"),
    (PAPER, "converted by us from their linear fit", "alpha provenance stated", "#151"),
    (REVIEW, "Round 2", "the external review round is recorded", "#151"),
    (PAPER, "academia.edu/170831186", "the systematics paper is cited, not left unverified", "#155"),
    (PAPER, "a0z_baryonic_systematics_check.py", "the coherence check is cited", "#155"),
    (PAPER, "UNTESTED", "the honest status is stated", "#155"),
    (REPORT, "Rusishvili", "the Report carries the systematics challenge", "#155"),
    (REPORT, "form-forced conditional on the live-horizon reading",
     "a0(z) tier synchronised with the workbook", "#161"),
    (REPORT, "2.3\u03c3 on a computed conversion budget",
     "Bottom Line carries the computed figure", "#161"),
    (REPORT, "the comparison with the Standard Model's own parameter inheritance is made once",
     "SM comparison argued once, referenced elsewhere", "#161"),
]

# --------------------------------------------------------------------------- C
# A shared number must appear in every artifact that discusses it, identically.
AGREE = [
    ("the computed tension", r"2\.3σ", [REPORT, PAPER,
                                            "physics-papers/predictions/ED_Master_Predictions_List.md",
                                            "physics-papers/predictions/Paper_101_FalsificationRegister.md",
                                            "physics-papers/predictions/22_Ways_to_Kill_Event_Density.md",
                                            "essays/Paper_SelectedFormalisms_I_Gravity.md"]),
    ("the refit central value", r"1\.15", [REPORT, PAPER,
                                           "physics-papers/predictions/ED_Master_Predictions_List.md"]),
    ("the survey's linear slope", r"1\.59", [PAPER,
                                             "physics-papers/predictions/ED_Master_Predictions_List.md"]),
]


def read(p):
    try:
        return io.open(p, encoding="utf-8").read()
    except IOError:
        return None


def rule(c="-", n=76):
    print(c * n)


# A document that narrates its own corrections must QUOTE the figures it
# superseded. Those are audit trail, not staleness. A match is exempt when one of
# these markers appears shortly before it.
CORRECTION_MARKERS = (
    r"earlier", r"supersed", r"in place of", r"previously", r"was wrong", r"asserted",
    r"superseded", r"corrected", r"replaced", r"no longer", r"used to", r"stale",
    r"special pleading", r"now quantified", r"which was asserted", r"not the survey", r"was the paper's soft joint",
)
CONTEXT = 220


def _live_hits(pat, text):
    """Matches that are NOT inside a sentence announcing a correction.

    The window looks BOTH WAYS. A correction is written either before the figure
    ("an earlier draft said X") or after it ("X -- and now quantified: Y"), and an
    exemption that only looks backwards misses the second form. That cost a false
    positive on the Report's own corrected flagship paragraph.
    """
    live = []
    for m in re.finditer(pat, text):
        window = text[max(0, m.start() - CONTEXT):m.end() + CONTEXT].lower()
        if any(re.search(k, window) for k in CORRECTION_MARKERS):
            continue
        live.append(m)
    return live


def check(texts):
    """Returns list of (class, file, message)."""
    out = []

    for path, pat, why, owner in STALE:
        t = texts.get(path)
        if t is None:
            out.append(("A", path, "FILE MISSING"))
            continue
        hits = _live_hits(pat, t)
        if hits:
            ln = t.count("\n", 0, hits[0].start()) + 1
            out.append(("A", path, "STALE (line %d): %s  [%s]" % (ln, why, owner)))

    for path, needle, why, owner in APPLIED:
        t = texts.get(path)
        if t is None:
            out.append(("B", path, "FILE MISSING"))
            continue
        if needle not in t:
            out.append(("B", path, "NOT APPLIED: %s (%r) [%s]" % (why, needle, owner)))

    for label, pat, files in AGREE:
        missing = [f for f in files if texts.get(f) and not re.search(pat, texts[f])]
        if missing:
            out.append(("C", ", ".join(os.path.basename(m) for m in missing),
                        "DISAGREES on %s (pattern %s absent)" % (label, pat)))
    return out


def freshness():
    out = []
    for md in (REPORT, PAPER):
        pdf = md[:-3] + ".pdf"
        if not os.path.exists(pdf):
            out.append(("D", pdf, "PDF MISSING"))
        elif os.path.getmtime(pdf) < os.path.getmtime(md):
            out.append(("D", pdf, "PDF IS OLDER THAN ITS SOURCE - rebuild"))

    # Desktop copy: compare CONTENT, not mtime. `git checkout` rewrites the working
    # tree and bumps mtimes without changing bytes, which made an identical copy
    # look stale on the first run of this script.
    desk = os.path.join(os.path.expanduser("~"), "Desktop", "ED_UnifiedFramework_Report.pdf")
    rp = REPORT[:-3] + ".pdf"
    if os.path.exists(desk) and os.path.exists(rp):
        import hashlib

        def h(p):
            with open(p, "rb") as fh:
                return hashlib.sha256(fh.read()).hexdigest()

        if h(desk) != h(rp):
            out.append(("D", "Desktop copy", "desktop PDF differs from the built one - re-copy"))

    try:
        import openpyxl
    except ImportError:
        out.append(("D", WORKBOOK, "openpyxl unavailable - workbook not checked"))
        return out
    if not os.path.exists(WORKBOOK):
        out.append(("D", WORKBOOK, "MISSING"))
        return out
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True)
    ws = wb["Ledger w Claims"]
    import collections
    live = collections.Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[3]:
            live[str(r[3]).strip()] += 1
    tc = wb["Tier Counts"]
    for row in tc.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        name, stored = str(row[0]), row[1]
        key = name.split("(")[0].strip()
        if key in live and isinstance(stored, int) and live[key] != stored:
            out.append(("D", "Tier Counts",
                        "%s: sheet says %d, rows say %d" % (key, stored, live[key])))
    return out


def main():
    selftest = "--selftest" in sys.argv
    paths = set(p for p, _, _, _ in STALE) | set(p for p, _, _, _ in APPLIED)
    for _, _, fs in AGREE:
        paths |= set(fs)
    texts = {p: read(p) for p in paths}

    print()
    rule("=")
    print("ARTIFACT CURRENCY CHECK - do the three documents say what I think they say?")
    rule("=")
    print("  checks: %d stale-patterns, %d applied-fixes, %d cross-doc agreements"
          % (len(STALE), len(APPLIED), len(AGREE)))
    print()

    findings = check(texts) + freshness()

    if selftest:
        rule()
        print("SELF-TEST - plant a known-stale string and confirm it is caught")
        rule()
        planted = dict(texts)
        planted[REPORT] = (planted[REPORT] or "") + "\nIt is a closed sector, but it still has edges:\n"
        got = [f for f in check(planted) if f[0] == "A" and "closed sector" in f[2]]
        if got:
            print("  PASS - the checker detects a planted stale string.")
        else:
            print("  FAIL - the checker did NOT detect a planted stale string.")
            findings.append(("SELFTEST", "-", "checker is not sensitive"))
        # and confirm a removed fix is caught
        planted2 = dict(texts)
        planted2[PAPER] = (planted2[PAPER] or "").replace("Magneticum", "XXX")
        got2 = [f for f in check(planted2) if f[0] == "B"]
        print("  %s - the checker detects a removed fix." % ("PASS" if got2 else "FAIL"))
        if not got2:
            findings.append(("SELFTEST", "-", "applied-check is not sensitive"))

        # The exemption must not swallow a genuine unqualified occurrence. Plant
        # one with no correction marker anywhere near it.
        planted3 = dict(texts)
        planted3[PAPER] = (planted3[PAPER] or "") + "\n\nThe tension is roughly 1\u20132\u03c3.\n"
        got3 = [f for f in check(planted3) if f[0] == "A" and "softening" in f[2]]
        print("  %s - the exemption does not swallow an unqualified stale claim."
              % ("PASS" if got3 else "FAIL"))
        if not got3:
            findings.append(("SELFTEST", "-", "exemption is too permissive"))
        print()

    rule()
    if not findings:
        print("RESULT: CLEAN")
        rule()
        print("  No stale text, every recorded fix present, shared numbers agree,")
        print("  PDFs newer than their sources, workbook counts match its own rows.")
        print()
        print("  This does NOT verify the claims are true. It verifies the documents")
        print("  say what the ledger says they say. Those are different jobs and only")
        print("  the second one can be automated.")
        print()
        return 0

    print("RESULT: %d FINDING(S)" % len(findings))
    rule()
    for cls, where, msg in findings:
        print("  [%s] %-46s %s" % (cls, os.path.basename(str(where))[:46], msg))
    print()
    return 1


if __name__ == "__main__":
    sys.exit(main())
