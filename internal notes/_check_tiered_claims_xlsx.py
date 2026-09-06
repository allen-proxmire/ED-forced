# -*- coding: utf-8 -*-
"""Cross-check ED_ItemizedTheory_TieredClaims.xlsx against the corpus.

The spreadsheet is AP's stated FIRST STOP for any tier-check.  On 2026-09-06 a
manual look found it internally sound but a 2026-09-04 snapshot: 115 of its 404
rows were last dated 2026-07-29, it names 92 of the corpus's 174 postulates, and
it misses the a_0 prior-art correction (gravity ledger #60) -- the single most
externally-sensitive claim in the corpus.

That is the same failure mode `_census_postulates.py` was built to stop for
postulate counts, and nothing played that role here.  This does.

FOUR CHECKS:

  C1  POSTULATE COVERAGE.  Named P-postulates in the corpus that the sheet
      never mentions.  The sheet's Postulate column cannot audit postulate load
      if it covers half of it.

  C2  DATE STALENESS.  Sheet rows whose paper has a gravity-ledger entry dated
      LATER than the row's own latest date.  A row that predates a finding
      about its own paper is a row that will mislead.

  C3  RESOLVED-BUT-OPEN.  Rows tagged Open/STALE whose paper appears in a
      ledger item containing a resolution word (CLOSED, SETTLED, RESOLVED,
      ANSWERED, ADOPTED).  Candidates for promotion.

  C4  DERIVED RESTING ON A POSTULATE.  Rows tagged Derived whose source paper
      declares a named P-postulate in its own text.  This is the mechanical
      form of the external audit's sharpest criticism -- "a postulate wearing a
      derivation's tag".  It is a FLAG, NOT A VERDICT: a paper may declare a
      postulate that is not load-bearing for the particular claim in the row.
      Each hit needs a human read.  What the check does is make the list finite
      and stop it being invisible.

Run: python "internal notes/_check_tiered_claims_xlsx.py"
     python "internal notes/_check_tiered_claims_xlsx.py" --full   (list every hit)
"""
import io
import os
import re
import subprocess
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

XLSX = "ED_ItemizedTheory_TieredClaims.xlsx"
LEDGERS = ["physics-papers/gravity/Gravity_TieredClaims_Ledger.md"]
FULL = "--full" in sys.argv
PNAME = re.compile(r"P-[A-Za-z0-9][A-Za-z0-9\-]{2,}")
DATE = re.compile(r"2026-\d\d-\d\d")
RESOLVED = re.compile(r"\b(CLOSED|SETTLED|RESOLVED|ANSWERED|ADOPTED|WITHDRAWN)\b")
FALSE_POSITIVES = {"P-CONSTRUCTION", "P-constructions", "P-definitional",
                   "P-imprint", "P-postulates"}


def census_names():
    out = subprocess.run([sys.executable, "internal notes/_census_postulates.py", "--list"],
                         capture_output=True, text=True, encoding="utf-8").stdout
    return set(re.findall(r"^  (P-[A-Za-z0-9][A-Za-z0-9\-]*)", out, re.M))


def load_sheet():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["Ledger w Claims"]
    hdr = [str(c) if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(hdr)}
    rows = []
    for n, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(r):
            continue
        rows.append({"n": n,
                     "paper": str(r[idx["Paper"]] or ""),
                     "claim": str(r[idx["Claim"]] or ""),
                     "tier": str(r[idx["Tier"]] or ""),
                     "status": str(r[idx["Status"]] or ""),
                     "postulate": str(r[idx["Postulate"]] or ""),
                     "all": " ".join(str(c) for c in r if c)})
    return rows


def ledger_items():
    """Ledger items keyed by the paper names they mention."""
    items = []
    for path in LEDGERS:
        if not os.path.exists(path):
            continue
        text = io.open(path, encoding="utf-8").read()
        for block in re.split(r"\n(?=\d+\. \*\*)", text):
            ds = DATE.findall(block)
            if not ds:
                continue
            items.append({"date": max(ds), "text": block,
                          "num": (re.match(r"(\d+)\.", block) or [None, "?"])[1]})
    return items


def paper_files():
    """Index every paper under several keys, because the sheet's Paper column
    uses bare numbers ("042"), short names ("GR-I") and full names alike.
    Resolving only on the full name silently reports zero hits -- which is a
    false all-clear on exactly the check C4 exists for."""
    out = {}
    for dp, _dn, fn in os.walk("physics-papers"):
        for f in fn:
            if not f.endswith(".md"):
                continue
            path = os.path.join(dp, f)
            stem = f[:-3]
            out.setdefault(stem, path)
            m = re.match(r"Paper_([0-9]+(?:_[0-9]+)?)", stem)
            if m:
                out.setdefault(m.group(1), path)               # "042"
                out.setdefault(m.group(1).lstrip("0"), path)    # "42"
            m2 = re.match(r"Paper_((?:GR|KM|MS|RQM|ED)[-_][A-Za-z0-9\-]+)", stem)
            if m2:
                out.setdefault(m2.group(1).replace("_", "-"), path)
    return out


def resolve(name, files):
    if name in files:
        return files[name]
    for k, v in files.items():
        if name and (name in k or k in name) and len(name) > 2:
            return v
    return None


def main():
    rows = load_sheet()
    cen = census_names()
    items = ledger_items()
    files = paper_files()

    sheet_names = set()
    for r in rows:
        sheet_names |= set(PNAME.findall(r["all"]))
    sheet_names -= FALSE_POSITIVES

    print("Tiered-claims spreadsheet cross-check")
    print("=" * 78)
    print("  sheet rows: %d      corpus postulates: %d      ledger items scanned: %d\n"
          % (len(rows), len(cen), len(items)))

    # ---- C1 ------------------------------------------------------------
    missing = sorted(cen - sheet_names)
    print("C1  POSTULATE COVERAGE")
    print("    in the corpus, never mentioned in the sheet : %d of %d"
          % (len(missing), len(cen)))
    show = missing if FULL else missing[:10]
    for m in show:
        print("       %s" % m)
    if not FULL and len(missing) > 10:
        print("       ... and %d more (--full)" % (len(missing) - 10))

    # ---- C2 ------------------------------------------------------------
    stale = []
    for r in rows:
        ds = DATE.findall(r["all"])
        rowdate = max(ds) if ds else None
        if not r["paper"]:
            continue
        later = [it for it in items
                 if r["paper"] in it["text"] and (rowdate is None or it["date"] > rowdate)]
        if later:
            stale.append((r, rowdate, later))
    print("\nC2  DATE STALENESS")
    print("    rows whose paper has a LATER ledger entry : %d" % len(stale))
    for r, rd, later in (stale if FULL else stale[:8]):
        print("       row %-4d %-28s row-date %s  <  ledger #%s (%s)"
              % (r["n"], r["paper"][:28], rd or "none", later[-1]["num"], later[-1]["date"]))
    if not FULL and len(stale) > 8:
        print("       ... and %d more (--full)" % (len(stale) - 8))

    # ---- C3 ------------------------------------------------------------
    promo = []
    for r in rows:
        if not r["paper"]:
            continue
        if "open" not in r["tier"].lower() and "stale" not in r["all"].upper():
            continue
        hits = [it for it in items if r["paper"] in it["text"] and RESOLVED.search(it["text"])]
        if hits:
            promo.append((r, hits[-1]))
    print("\nC3  RESOLVED-BUT-STILL-OPEN")
    print("    open/stale rows whose paper has a resolving ledger item : %d" % len(promo))
    for r, it in (promo if FULL else promo[:8]):
        print("       row %-4d %-28s -> ledger #%s (%s)"
              % (r["n"], r["paper"][:28], it["num"], it["date"]))
    if not FULL and len(promo) > 8:
        print("       ... and %d more (--full)" % (len(promo) - 8))

    # ---- C4 ------------------------------------------------------------
    flagged = []
    for r in rows:
        if not r["tier"].lower().startswith("derived"):
            continue
        f = resolve(r["paper"], files)
        if not f:
            continue
        text = io.open(f, encoding="utf-8", errors="replace").read()
        decl = set(PNAME.findall(text)) - FALSE_POSITIVES
        decl &= cen
        if decl:
            flagged.append((r, sorted(decl)))
    derived_total = sum(1 for r in rows if r["tier"].lower().startswith("derived"))
    print("\nC4  DERIVED ROWS WHOSE PAPER DECLARES A NAMED POSTULATE")
    print("    (a FLAG, not a verdict -- the postulate may not be load-bearing")
    print("     for this particular claim; each hit needs a human read)")
    print("    flagged : %d of %d Derived rows" % (len(flagged), derived_total))
    for r, d in (flagged if FULL else flagged[:8]):
        print("       row %-4d %-26s declares %s"
              % (r["n"], r["paper"][:26], ", ".join(d[:3]) + (" ..." if len(d) > 3 else "")))
    if not FULL and len(flagged) > 8:
        print("       ... and %d more (--full)" % (len(flagged) - 8))

    print("""
WHAT TO DO WITH THIS

  C1 and C2 are mechanical and should go to zero: add the missing postulates,
  re-date the rows whose papers have moved.

  C3 is free improvement -- rows that can be promoted out of Open.

  C4 is the one that needs judgement, and it is the mechanical form of the
  external audit's sharpest criticism: "a postulate wearing a derivation's
  tag".  A hit does NOT mean the row is mislabelled.  It means the row's paper
  contains a named postulate and somebody should check whether THIS claim
  depends on it.  The value is that the list is finite and visible instead of
  being an open-ended worry.
""")
    return 1 if (missing or stale) else 0


if __name__ == "__main__":
    sys.exit(main())
