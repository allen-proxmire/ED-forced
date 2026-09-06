# -*- coding: utf-8 -*-
"""C1, triaged: which UNMENTIONED postulates actually matter?

The C1 inside `_check_tiered_claims_xlsx.py` reports a bare count -- 87 of 174
corpus postulates never named in the workbook -- and a bare count is not a work
list.  `POSTULATE_BASIS.md` (research target #20, closed 2026-09-04) already
established that the corpus's postulates form a BREADTH LADDER, and that 92 of
the 174 appear in exactly one paper: local modelling choices scoped to a single
derivation, like a gauge choice.  A one-paper postulate missing from a
corpus-wide claim catalogue is not obviously a defect.

So this crosses the unmentioned set against the ladder and reports by rung:

  CROSS-CUTTING (>=4 papers)  a standing commitment of the framework, absent
                              from the catalogue.  A real gap.
  RECURRING (2-3 papers)      outlives the derivation that introduced it.
                              Worth a row.
  LOCAL (1 paper)             scoped to one derivation.  Absence is defensible;
                              listing all of them would bury the others.

It also separates a fourth class the bare count hides:

  COVERED BY A ROW  the postulate is not named in the sheet, but its PAPER has
                    a row, so the claim it supports is catalogued even though
                    the commitment is not.  Different, milder defect.

Run: python "internal notes/_check_c1_coverage.py"
     python "internal notes/_check_c1_coverage.py" --full
"""
import io
import os
import re
import subprocess
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)

XLSX = next((a for a in sys.argv[1:] if a.endswith(".xlsx")),
            "ED_ItemizedTheory_TieredClaims_v2.xlsx")
FULL = "--full" in sys.argv
PNAME = re.compile(r"P-[A-Za-z0-9][A-Za-z0-9\-]{2,}")
FALSE_POSITIVES = {"P-CONSTRUCTION", "P-constructions", "P-definitional",
                   "P-imprint", "P-postulates"}


def census():
    out = subprocess.run([sys.executable, "internal notes/_census_postulates.py", "--list"],
                         capture_output=True, text=True, encoding="utf-8").stdout
    return set(re.findall(r"^  (P-[A-Za-z0-9][A-Za-z0-9\-]*)", out, re.M))


def where_declared():
    """postulate -> set of paper files that name it."""
    d = defaultdict(set)
    for dp, _dn, fn in os.walk("physics-papers"):
        for f in fn:
            if not f.endswith(".md") or f.endswith("_TieredClaims_Ledger.md"):
                continue
            p = os.path.join(dp, f)
            for m in set(PNAME.findall(io.open(p, encoding="utf-8", errors="replace").read())):
                if m not in FALSE_POSITIVES:
                    d[m].add(p)
    return d


def main():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    named, papers_with_rows = set(), set()
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            blob = " ".join(str(c) for c in row if c)
            named |= set(PNAME.findall(blob))
            if sheet == "Ledger w Claims" and row and row[1]:
                papers_with_rows.add(str(row[1]).strip())
    named -= FALSE_POSITIVES

    cen = census()
    decl = where_declared()
    missing = sorted(cen - named)

    rungs = {"CROSS-CUTTING (>=4 papers)": [], "RECURRING (2-3 papers)": [],
             "LOCAL (1 paper)": []}
    for m in missing:
        n = len(decl.get(m, ()))
        key = ("CROSS-CUTTING (>=4 papers)" if n >= 4
               else "RECURRING (2-3 papers)" if n >= 2 else "LOCAL (1 paper)")
        rungs[key].append((m, n, sorted(decl.get(m, ()))))

    print("C1 TRIAGED - unmentioned postulates by breadth")
    print("=" * 78)
    print("  workbook: %s" % XLSX)
    print("  corpus postulates: %d      named in the workbook: %d      unmentioned: %d\n"
          % (len(cen), len(cen & named), len(missing)))

    for key in ("CROSS-CUTTING (>=4 papers)", "RECURRING (2-3 papers)", "LOCAL (1 paper)"):
        got = rungs[key]
        print("  %-30s %d" % (key, len(got)))
    print()

    for key in ("CROSS-CUTTING (>=4 papers)", "RECURRING (2-3 papers)"):
        got = rungs[key]
        if not got:
            continue
        print("  --- %s ---" % key)
        for m, n, files in got:
            covered = any(os.path.basename(f)[:-3] in " ".join(papers_with_rows)
                          or any(k and k in os.path.basename(f) for k in papers_with_rows)
                          for f in files)
            print("    %-38s %d paper(s)%s" % (m, n, "   [paper HAS a row]" if covered else ""))
            if FULL:
                for f in files[:5]:
                    print("        %s" % f)
        print()

    if FULL and rungs["LOCAL (1 paper)"]:
        print("  --- LOCAL (1 paper) ---")
        for m, n, files in rungs["LOCAL (1 paper)"]:
            print("    %-38s %s" % (m, os.path.basename(files[0]) if files else "-"))

    print("""
  READ THIS AS A LADDER, NOT A COUNT.  A postulate used in ONE paper is a local
  modelling choice; its absence from a corpus-wide catalogue is defensible and
  listing all of them would bury the ones that matter.  The CROSS-CUTTING rung
  is the real gap: a standing commitment of the framework with no row.
""")
    return 0


if __name__ == "__main__":
    sys.exit(main())
